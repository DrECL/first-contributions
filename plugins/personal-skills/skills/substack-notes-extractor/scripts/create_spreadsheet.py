#!/usr/bin/env python3
"""Create a formatted .xlsx spreadsheet from Substack Notes data."""
import argparse
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_spreadsheet(posts_data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Substack Notes"

    headers = ["Date", "Text of Post", "Likes (Hearts)", "Comments", "Restacks", "Link of Post"]
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_font = Font(name="Arial", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    alt_fill = PatternFill("solid", fgColor="F2F7FB")

    for r, post in enumerate(posts_data, 2):
        fill = alt_fill if r % 2 == 0 else PatternFill()

        ws.cell(row=r, column=1, value=post["date"]).font = data_font
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).border = thin_border

        ws.cell(row=r, column=2, value=post["text"]).font = data_font
        ws.cell(row=r, column=2).alignment = wrap_align
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=2).border = thin_border

        for col, key in [(3, "likes"), (4, "comments"), (5, "restacks")]:
            c = ws.cell(row=r, column=col, value=int(post[key]))
            c.font = data_font
            c.alignment = Alignment(horizontal="center", vertical="top")
            c.fill = fill
            c.border = thin_border

        ws.cell(row=r, column=6, value=post["link"]).font = link_font
        ws.cell(row=r, column=6).alignment = Alignment(vertical="top")
        ws.cell(row=r, column=6).fill = fill
        ws.cell(row=r, column=6).border = thin_border

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 55

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(posts_data) + 1}"

    wb.save(output_path)
    print(f"Saved {len(posts_data)} posts to {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    parser.add_argument("--data", required=True, help="JSON array of post objects")
    args = parser.parse_args()

    posts = json.loads(args.data)
    create_spreadsheet(posts, args.output)
